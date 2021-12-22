# OK...my contact form in python, to be used as a template for any form to be
# submitted to a database of some sort...

# import modules to use excel and some web-based driver, tkinter in this case
# and an input sanitizer

import tkinter as tk
import openpyxl as ox
from tkinter import scrolledtext
import bleach

# open pre-existing excel file...
file_path = 'C:\\Users\\dad\\Desktop\\Contact Us\\excel.xlsx'
xlwb = ox.load_workbook(file_path)

'''
# ...or create one if neccessary
# there are several ways to do this, lets try this one
# so just import openpyxl as done above...

import os

if not os.path.isdir('C:\\Users\\dad\\Desktop\\Contact Us\\'):
    os.makedirs('C:\\Users\\dad\\Desktop\\Contact Us\\')

file_path = 'C:\\Users\\dad\\Desktop\\Contact Us\\excel.xlsx'

if not os.path.isfile(file_path):
    xlwb = ox.Workbook(file_path)
    xlwb.save(filename = file_path)
    
xlwb = ox.load_workbook(file_path)
'''

# create worksheet within workbook
xlsheet = xlwb.active

# define worksheet features...

# set column widths for worksheet
def excel():
    xlsheet.column_dimensions['A'].width = 40
    xlsheet.column_dimensions['B'].width = 40
    xlsheet.column_dimensions['C'].width = 20
    xlsheet.column_dimensions['D'].width = 30
    xlsheet.column_dimensions['E'].width = 20

    # write given headers to first row of worksheet
    xlsheet.cell(row=1, column=1).value = 'Name'
    xlsheet.cell(row=1, column=2).value = 'Email'
    xlsheet.cell(row=1, column=3).value = 'Area Code'
    xlsheet.cell(row=1, column=4).value = 'Phone Number'
    xlsheet.cell(row=1, column=5).value = 'Message'


# functions to set focus on fields...
def focus1(event):
    name_field.focus_set()


def focus2(event):
    email_field.focus_set()


def focus3(event):
    area_code_field.focus_set()


def focus4(event):
    phone_no_field.focus_set()


def focus5(event):
    message_field.focus_set()


# clear all fields function (after submission)
def clear():

    name_field.delete(0, 'end')
    email_field.delete(0, 'end')
    area_code_field.delete(0, 'end')
    phone_no_field.delete(0, 'end')

    # slightly different for text boxes
    message_field.delete(1.0, 'end')


# Function to take info from form and write to file

def insert():
    # if no entries...
    if (name_field.get() == '' or
        email_field.get() == '' or
        
        # text boxes never truly empty, always has '\n'
        # but not displayed
        message_field.get(1.0) == '\n'
        ):
       
        print('Please complete boxes with * ...no message recorded!')

    else:
        current_row = xlsheet.max_row

        # get method returns current text
        # as string which we write into
        # worksheet at particular location
        # after sanitizing

        xlsheet.cell(row=current_row + 1,
                     column=1).value = bleach.clean(name_field.get()[0:70])
        xlsheet.cell(row=current_row + 1,
                     column=2).value = bleach.clean(email_field.get()[0:100])
        xlsheet.cell(row=current_row + 1,
                     column=3).value = area_code_field.get()[0:6]
        xlsheet.cell(row=current_row + 1,
                     column=4).value = phone_no_field.get()[0:7]
        # text boxes are a little different...
        xlsheet.cell(row=current_row + 1,
                     column=5).value = bleach.clean(message_field.get(1.0, 'end')[0:250])

        # save to worksheet file
        xlwb.save(file_path)

        # reset focus on the name_field box
        name_field.focus_set()

        # clear form after submission
        clear()


# Driver code
if __name__ == '__main__':

    # create window in tkinter
    window = tk.Tk()
    window.title('Contact Us!!!')

    
    # create frames and labels within for GUI objects
    frame1 = tk.Frame(master=window, width=550, height=80, borderwidth=1,
                      bg='orange')
    frame1.pack()

    label1 = tk.Label(master=frame1, text='Contact Us', bg='orange',
                      fg='white', font=('New Times Roman', 20))
    label1.place(x=25, y=20)

    frame2 = tk.Frame(master=window, width=550, height=70, borderwidth=1)
    frame2.pack()

    label2 = tk.Label(master=frame2, text='Name', font=('New Times Roman', 13),
                      fg='gray')
    label2.place(x=25, y=10)

    label2a = tk.Label(master=frame2, text='*', font=('New Times Roman', 13),
                       fg='red')
    label2a.place(x=75, y=10)

    name_field = tk.Entry(master=frame2, font=('New Times Roman', 15),
                          width=45, bd=0)
    name_field.place(x=25, y=40)

    frame3 = tk.Frame(master=window, width=550, height=70, borderwidth=1)
    frame3.pack()

    label3 = tk.Label(master=frame3, text='Email',
                      font=('New Times Roman', 13), fg='gray')
    label3.place(x=25, y=10)

    label3a = tk.Label(master=frame3, text='*', font=('New Times Roman', 13),
                       fg='red')
    label3a.place(x=75, y=10)

    email_field = tk.Entry(master=frame3, font=('New Times Roman', 15),
                           width=45, bd=0)
    email_field.place(x=25, y=40)

    frame4 = tk.Frame(master=window, width=550, height=90, borderwidth=1)
    frame4.pack()

    label4 = tk.Label(master=frame4, text='Phone Number',
                      font=('New Times Roman', 13), fg='gray')
    label4.place(x=25, y=10)

    area_code_field = tk.Entry(master=frame4, font=('New Times Roman', 15),
                               width=15, bd=0)
    area_code_field.place(x=25, y=40)

    phone_no_field = tk.Entry(master=frame4, font=('New Times Roman', 15),
                              width=28, bd=0)
    phone_no_field.place(x=215, y=40)

    label4a = tk.Label(master=frame4, text='Area Code',
                       font=('New Times Roman', 10), fg='light gray')
    label4a.place(x=25, y=67)

    label4b = tk.Label(master=frame4, text='Phone Number',
                       font=('New Times Roman', 10), fg='light gray')
    label4b.place(x=215, y=67)

    frame5 = tk.Frame(master=window, width=550, height=150, borderwidth=1)
    frame5.pack()

    label5 = tk.Label(master=frame5, text='Message',
                      font=('New Times Roman', 13), fg='gray')
    label5.place(x=25, y=10)

    label5a = tk.Label(master=frame5, text='*', font=('New Times Roman', 13),
                       fg='red')
    label5a.place(x=95, y=10)

    # text boxes are a little different...
    message_field = scrolledtext.ScrolledText(master=frame5,
                                              font=('New Times Roman', 15),
                                              width=45, height=4, bd=0,
                                              wrap=tk.WORD)
    message_field.place(x=25, y=40)

    frame6 = tk.Frame(master=window, width=550, height=80, borderwidth=1)
    frame6.pack()

    # submit button
    button1 = tk.Button(master=frame6, text='Submit', bg='orange', fg='white',
                        font=('New Times Roman', 11), command=insert, bd=0,
                        width=11, height=2)
    button1.place(x=410, y=8)

    # call up worksheet
    excel()

    window.mainloop()
